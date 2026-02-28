"""
generates a styled Excel report for non-technical stakeholders.
6 sheets: executive summary, fraud details, stats, method comparison, bank performance, recommendations.
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def style_header(ws, row=1, cols=10):
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border


def create_excel_report(df, comparison_df, hyp_results_df, output_path):
    wb = Workbook()

    # sheet 1: high-level summary for managers
    ws1 = wb.active
    ws1.title = 'Executive Summary'
    ws1['A1'] = 'UPI Transaction Fraud Analysis Report'
    ws1['A1'].font = Font(size=16, bold=True, color='1F4E79')
    ws1['A3'] = 'Report Generated:'
    ws1['B3'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')

    kpis = [
        ('Total Transactions', f"{len(df):,}"),
        ('Total Transaction Value', f"Rs {df['amount'].sum():,.0f}"),
        ('Average Transaction Value', f"Rs {df['amount'].mean():,.2f}"),
        ('Median Transaction Value', f"Rs {df['amount'].median():,.2f}"),
        ('Fraud Transactions Detected', f"{df['is_fraud'].sum():,}"),
        ('Fraud Rate', f"{df['is_fraud'].mean()*100:.2f}%"),
        ('Estimated Fraud Value', f"Rs {df[df['is_fraud']==1]['amount'].sum():,.0f}"),
        ('Unique Senders', f"{df['sender_upi_id'].nunique():,}"),
        ('Date Range', f"{df['timestamp'].min().strftime('%Y-%m-%d')} to {df['timestamp'].max().strftime('%Y-%m-%d')}"),
    ]
    for i, (label, value) in enumerate(kpis):
        ws1.cell(row=5+i, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=5+i, column=2, value=value)
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 25

    # sheet 2: top 500 fraud transactions sorted by amount
    ws2 = wb.create_sheet('Fraud Analysis')
    fraud_df = df[df['is_fraud'] == 1][['transaction_id', 'timestamp', 'sender_upi_id',
                                         'amount', 'fraud_type', 'city', 'transaction_type']].head(500)
    fraud_df = fraud_df.sort_values('amount', ascending=False)
    for r_idx, row in enumerate(dataframe_to_rows(fraud_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws2.cell(row=r_idx, column=c_idx, value=value)
    style_header(ws2, row=1, cols=len(fraud_df.columns))

    # sheet 3: hypothesis test results
    ws3 = wb.create_sheet('Statistical Results')
    ws3['A1'] = 'Hypothesis Testing Results'
    ws3['A1'].font = Font(size=14, bold=True, color='1F4E79')
    if hyp_results_df is not None and len(hyp_results_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(hyp_results_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                ws3.cell(row=r_idx, column=c_idx, value=str(value) if not isinstance(value, (int, float)) else value)
        style_header(ws3, row=3, cols=len(hyp_results_df.columns))

    # sheet 4: precision/recall/f1 for each detection method
    ws4 = wb.create_sheet('Detection Methods')
    ws4['A1'] = 'Fraud Detection Method Comparison'
    ws4['A1'].font = Font(size=14, bold=True, color='1F4E79')
    if comparison_df is not None and len(comparison_df) > 0:
        for r_idx, row in enumerate(dataframe_to_rows(comparison_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = ws4.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, float):
                    cell.number_format = '0.00%'
        style_header(ws4, row=3, cols=len(comparison_df.columns))

    # sheet 5: bank-level aggregates
    ws5 = wb.create_sheet('Bank Performance')
    bank_stats = df.groupby('sender_bank').agg(
        total_txns=('transaction_id', 'count'),
        total_value=('amount', 'sum'),
        avg_amount=('amount', 'mean'),
        fraud_count=('is_fraud', 'sum'),
        fail_count=('status', lambda x: (x == 'FAILED').sum())
    ).reset_index()
    bank_stats['fraud_rate'] = bank_stats['fraud_count'] / bank_stats['total_txns']
    bank_stats['fail_rate'] = bank_stats['fail_count'] / bank_stats['total_txns']
    bank_stats = bank_stats.sort_values('total_txns', ascending=False)
    for r_idx, row in enumerate(dataframe_to_rows(bank_stats.round(4), index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws5.cell(row=r_idx, column=c_idx, value=value)
    style_header(ws5, row=1, cols=len(bank_stats.columns))

    # sheet 6: practical recommendations based on what we found
    ws6 = wb.create_sheet('Recommendations')
    ws6['A1'] = 'Data-Driven Recommendations'
    ws6['A1'].font = Font(size=14, bold=True, color='1F4E79')
    recs = [
        ('1. Implement velocity checks',
         'Flag accounts exceeding 10 transactions/hour. Catches rapid-fire fraud.'),
        ('2. Add threshold monitoring',
         'Watch transactions between Rs 9,000-9,999. Structuring is common.'),
        ('3. Night-time transaction limits',
         'Extra verification for txns above Rs 5,000 between 1 AM and 5 AM.'),
        ('4. New account cooling period',
         'Limit frequency for UPI IDs less than 7 days old.'),
        ('5. Geographic validation',
         'Flag same user in two distant cities within short timeframe.'),
        ('6. Use ensemble detection',
         'Combine Z-Score + Isolation Forest + Rules. 2/3 agreement works best.'),
    ]
    for i, (title, desc) in enumerate(recs):
        row = 3 + i * 3
        ws6.cell(row=row, column=1, value=title).font = Font(bold=True, size=11)
        ws6.cell(row=row+1, column=1, value=desc)
    ws6.column_dimensions['A'].width = 80

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)
    print(f"report saved: {output_path}")


if __name__ == '__main__':
    base_dir = os.path.dirname(os.path.dirname(__file__))
    processed_path = os.path.join(base_dir, 'data', 'processed', 'upi_transactions_flagged.csv')

    if os.path.exists(processed_path):
        df = pd.read_csv(processed_path, parse_dates=['timestamp'])
        comp_path = os.path.join(base_dir, 'data', 'processed', 'method_comparison.csv')
        hyp_path = os.path.join(base_dir, 'data', 'processed', 'hypothesis_tests.csv')
        comp_df = pd.read_csv(comp_path) if os.path.exists(comp_path) else None
        hyp_df = pd.read_csv(hyp_path) if os.path.exists(hyp_path) else None
        output_path = os.path.join(base_dir, 'reports', 'fraud_report.xlsx')
        create_excel_report(df, comp_df, hyp_df, output_path)
    else:
        print("flagged data not found. run fraud_detector.py first.")
